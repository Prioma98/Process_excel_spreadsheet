import openpyxl as xl
from openpyxl.chart import BarChart,Reference

#Loading the excel sheet in a variable 
wb=xl.load_workbook('Process_excel_spreadsheet\emp_details_edited.xlsx')
sheet=wb['Sheet1']

#all the required variables
listt,previous_salary_list=[],[]
corrected_salary=0
symbol='$,\,'
final_list=[]

#storing the hike percentage column of the sheet in the list named listt for later computation
for row_num in range(2,sheet.max_row+1):
    hike_percentage=sheet.cell(row=row_num,column=5).value
    listt.append(hike_percentage)

#storing the salary column of the sheet in the list named previous_salary_list
for rows in range(2,sheet.max_row+1):
    previous_salary=sheet.cell(row=rows,column=3).value
    previous_salary_list.append(previous_salary)

#getting rid of the symbols like:$ ,
for element in previous_salary_list:
    temp = ""
    for sym in element:
        if sym not in symbol:
            temp += sym
    final_list.append(temp)#final salary list with float number on which any computation can be done

#final computation on each salary of the column that is the hike percentage is deducted from the salary and rs 100 is added which is the final salary 
for el in range(0,len(final_list)):
    corrected_salary=float(final_list[el])-(float(final_list[el])*listt[el])
    corrected_salary_cell=sheet.cell(row=el+2,column=7)
    corrected_salary_cell.value=corrected_salary

#for making the Barchart of the corrected salary
values_to_show=Reference(sheet,
min_row=2,
max_row=sheet.max_row,
min_col=7,
max_col=7)
chart=BarChart()
chart.add_data(values_to_show)
sheet.add_chart(chart,'h2')

#finally it is stored in a different file named emp_details
wb.save('Process_excel_spreadsheet\emp1_details.xlsx')
